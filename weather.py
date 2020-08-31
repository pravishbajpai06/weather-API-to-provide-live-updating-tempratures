#Here is the code to provide temprature updates of a city by using openweather API 
import requests
from pprint import pprint

def weather_data(query):
    res=requests.get('https://api.openweathermap.org/data/2.5/weather?'+query+'&appid=c8b92668120f071c0ee81409f2f1ad54'); #openweather site is called here. 
    return res.json();
values = []
def print_weather(result,city):
    temp = "{}'s temperature: {}°C ".format(city,result['main']['temp'])
    values.append(temp)
    wind_speed = "Wind speed: {} m/s".format(result['wind']['speed'])
    values.append(wind_speed)
    wea_forecast = "Description: {}".format(result['weather'][0]['description'])
    values.append(wea_forecast)
    conditions = "Weather: {}".format(result['weather'][0]['main'])
    values.append(conditions)
    print(temp)
    print(wind_speed)
    print(wea_forecast)
    print(conditions)

city=input('Enter the city:')# Enter the name of the city
print()
try:
    query='q='+city;
    w_data=weather_data(query);
    print_weather(w_data, city)
    print()
except:
    print('City name not found...')
    
val_real = [city]
for i in range(len(values)):
    for j in range(len(values[i])):
        if(values[i][j] == ":"):
            flag = j
    x = values[i][flag+2:len(values[i])]
    val_real.append(x)

Cols = ["A","B","C","D","E"]
initials = ["City","Temperature","Wind Speed","Forecast","Conditions"]

from openpyxl import Workbook#imported pyxl library to update temprture reports to the excel sheet

workbook = Workbook()
file = workbook.active

xlsx_file = Path('updated.xlsx')
  wb_obj = openpyxl.load_workbook(xlsx_file) 
  sheet = wb_obj.active

  # For initial update. When True all the rows will be updated, whether or not "Update(0/1)" column is 1 or 0.
  initial = True

  # To store Total Updated rows in last attempt. Will be used to terminate the program, In case no rows are being updated.
  tupdated = 0

  # Number of Rows sheet.
  n = sheet.max_row

  # Empty List used for terminator thread.
  flag =[]

  # Thread Launch
  _thread.start_new_thread(terminator, (flag,))

  # Run until the flag list is Empty.
  while not flag:
    tupdated = 0

    # Run for Each Row
    for i in range(1, n):

      # if flag is empty and either this is an initial update or "Update(0/1)" column in sheet is 1
      if ((sheet.cell(row = i+1, column = 5).value) == 1 or initial) and not flag:
        
        tupdated+=1

        # Extract City and get the values for it
        city = sheet.cell(row = i+1, column = 1).value
        values = get_Value(city)

        # If temperature in celsius is required. 
        if (sheet.cell(row = i+1, column = 4).value) == 'C':

          # Update Humidity
          sheet.cell(row = i+1, column = 3).value =  values[1]
          # Update Temperature
          sheet.cell(row = i+1, column = 2).value =  values[2]
          # Save to file
          wb_obj.save("Values.xlsx") 

          print(f'Location: \t{values[0]}')
          print(f'Humidity: \t{values[1]}')
          print(f'Temperature: \t{values[2]}°C')
          print("----------------------------------------------------")
        
        else:
          # Update Humidity
          sheet.cell(row = i+1, column = 3).value =  values[1]
          # Update Temperature
          sheet.cell(row = i+1, column = 2).value =  toFah(values[2])
          # Save to file
          wb_obj.save("Values.xlsx")

          print(f'Location: \t{values[0]}')
          print(f'Humidity: \t{values[1]}')
          print(f'Temperature: \t{toFah(values[2])}°F')
          print("----------------------------------------------------")

      time.sleep(1)
    initial = False

    if tupdated == 0:
      print("Nothing to update. Exiting!!")
      break
    


if __name__ == "__main__":
    main()



